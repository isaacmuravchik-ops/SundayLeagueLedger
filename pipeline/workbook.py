"""
Workbook parser — downloads the stats .xlsx attachments from Gmail and extracts:
  - Canonical player list  → pipeline/canonical_players.json
  - Official +/− standings → pipeline/official_standings_<year>.json

The workbook is the source of truth for player identity: if two names both
appear in it, they are distinct people and must never be merged.

Run after ingest.py and classify.py.
"""

import base64
import json
import os
import re
import tempfile
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

load_dotenv()

RAW_DIR = Path(__file__).parent / "raw"
CLASSIFIED = Path(__file__).parent / "classified.json"
TOKEN_PATH = Path(__file__).parent / "token.json"
CANONICAL_OUT = Path(__file__).parent / "canonical_players.json"

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
YEAR_RE = re.compile(r"soccer\s+(\d{4})", re.IGNORECASE)


def _get_service():
    creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
    return build("gmail", "v1", credentials=creds)


def _download_attachment(service, msg_id: str, attachment_id: str) -> bytes:
    result = service.users().messages().attachments().get(
        userId="me", messageId=msg_id, id=attachment_id
    ).execute()
    return base64.urlsafe_b64decode(result["data"] + "==")


def _extract_year(filename: str) -> int | None:
    m = YEAR_RE.search(filename)
    return int(m.group(1)) if m else None


def _parse_workbook(wb_bytes: bytes, year: int) -> dict:
    """
    Parse a soccer stats xlsx. Returns:
      { "year": int, "players": [str], "standings": [{nickname, pm, g, gp}] }

    The workbook layout varies slightly year to year. We look for a sheet with
    player names in column A and numeric data in adjacent columns. We find the
    header row by looking for common column names ('+/-', 'Goals', 'GP', etc.).
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(wb_bytes)
        tmp_path = f.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
    finally:
        os.unlink(tmp_path)

    # Pick the first sheet (or the one most likely to be the standings)
    sheet = wb.active

    # Scan for the header row — look for a row containing '+/-' or 'wins'
    header_row = None
    col_map = {}
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if any(x in row_lower for x in ["+/-", "wins", "losses", "goals", "gp", "games"]):
            header_row = row_idx
            for ci, val in enumerate(row_lower):
                if "+/-" in val or "plus" in val:
                    col_map["pm"] = ci
                elif val in ("wins", "w"):
                    col_map["w"] = ci
                elif val in ("losses", "l"):
                    col_map["l"] = ci
                elif val in ("goals", "g", "goal"):
                    col_map["g"] = ci
                elif val in ("gp", "games played", "games"):
                    col_map["gp"] = ci
            break

    players = []
    standings = []

    for row in sheet.iter_rows(min_row=(header_row or 1) + 1, values_only=True):
        name = row[0]
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()
        players.append(name)

        entry = {"nickname": name}
        for stat, ci in col_map.items():
            val = row[ci] if ci < len(row) else None
            try:
                entry[stat] = int(val) if val is not None else None
            except (TypeError, ValueError):
                entry[stat] = None
        standings.append(entry)

    return {"year": year, "players": players, "standings": standings}


def run():
    if not CLASSIFIED.exists():
        print("[workbook] classified.json not found — run classify.py first")
        return

    classified = json.loads(CLASSIFIED.read_text())
    stats_msgs = {mid: info for mid, info in classified.items() if info["kind"] == "stats"}
    print(f"[workbook] {len(stats_msgs)} stats emails found")

    if not stats_msgs:
        print("[workbook] nothing to do")
        return

    service = _get_service()

    all_players: set[str] = set()
    by_year: dict[int, dict] = {}

    for mid in stats_msgs:
        raw_path = RAW_DIR / f"{mid}.json"
        if not raw_path.exists():
            continue
        msg = json.loads(raw_path.read_text())
        for att in msg.get("attachments", []):
            if not att.get("filename", "").endswith(".xlsx"):
                continue
            year = _extract_year(att["filename"]) or 2026
            print(f"[workbook] downloading {att['filename']} ({year})")
            wb_bytes = _download_attachment(service, mid, att["attachment_id"])
            parsed = _parse_workbook(wb_bytes, year)
            all_players.update(parsed["players"])
            by_year[year] = parsed
            out = Path(__file__).parent / f"official_standings_{year}.json"
            out.write_text(json.dumps(parsed, indent=2, ensure_ascii=False))
            print(f"[workbook] wrote {out.name} ({len(parsed['players'])} players)")

    # Canonical player list = union of all workbook names across all years
    canonical = sorted(all_players)
    CANONICAL_OUT.write_text(json.dumps(canonical, indent=2, ensure_ascii=False))
    print(f"[workbook] canonical_players.json: {len(canonical)} players")


if __name__ == "__main__":
    run()
